# -*- coding: utf-8 -*-
"""
Created on Thu Dec 20 11:04:41 2018

@author: leolqli
"""
import openpyxl

excel1 = openpyxl.load_workbook('./data/0a053f8ebf74e698dbb716e409b23c07.xlsx')

sheet1 = excel1.active

#check the merge_cells
#print sheet1.merged_cell_ranges

print sheet1.max_row

#the header must in (10) rows
N = 10


#store the number of columns
col_num = []
#iterater each row
for row in sheet1.iter_rows(max_row = N):
    num_in_each_row = 0
    for cell in row:
        if(cell.value != None):
            num_in_each_row += 1
    col_num.append(num_in_each_row)
    